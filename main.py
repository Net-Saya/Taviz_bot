from typing import Final
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
)
from datetime import datetime, timedelta
from openpyxl import Workbook
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import json
import os

# === Налаштовунання ===
TOKEN: Final = ""

# === Дані ===
user_stats = {}   # chat_id -> month -> user_id -> {"name": username, "text": int, "photo": int, "sticker": int, "gif": int, "total": int}
chat_names = {}   # chat_id -> назва групи
auto_report_config = {}  # chat_id -> {"admin_id": int, "admin_name": str}
CONFIG_FILE = "auto_report_config.json"


# === Завантаження конфігурації ===
def load_config():
    global auto_report_config
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                auto_report_config = json.load(f)
        except Exception as e:
            print(f"[ERROR] Помилка завантаження конфігурації: {e}")
            auto_report_config = {}


# === Збереження конфігурації ===
def save_config():
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(auto_report_config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[ERROR] Помилка збереження конфігурації: {e}")


# === /start ===
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Привіт! Я вважаю повідомлення у групах📊\n"
     
    )


# === /startrecord - Активація автоматичної відправки статистики ===
async def startrecord_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Дозволяє адміністратору активувати автоматичну відправку Excel статистики кожен місяць"""
    
    # Команда доступна тільки в групах
    if update.effective_chat.type == "private":
        await update.message.reply_text(
            "❌ Ця команда доступна тільки в групах. "
            "Напишіть /startrecord у групі, де ви адміністратор."
        )
        return
    
    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    chat_title = update.effective_chat.title
    
    # Перевіряємо, чи користувач адміністратор групи
    is_admin = await is_group_admin(user_id, chat_id, context)
    if not is_admin:
        await update.message.reply_text(
            "❌ Ви повинні бути адміністратором групи для активації автоматичної відправки статистики."
        )
        return
    
    # Зберігаємо конфігурацію
    chat_id_str = str(chat_id)
    auto_report_config[chat_id_str] = {
        "admin_id": user_id,
        "admin_name": update.effective_user.full_name,
        "chat_title": chat_title,
        "enabled": True
    }
    save_config()
    
    await update.message.reply_text(
        "✅ Успішно активовано! Розраховуюся що я буду сенндити вам Excel звіт щомісяця об першому числу о 9:00."
    )
    
    print(f"[LOG] /startrecord активована для групи {chat_id} ({chat_title}) адміністратором {user_id}")


# === /stoprecord - Деактивація автоматичної відправки статистики ===
async def stoprecord_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Деактивує автоматичну відправку Excel статистики"""
    
    # Команда доступна тільки в групах
    if update.effective_chat.type == "private":
        await update.message.reply_text(
            "❌ Ця команда доступна тільки в групах."
        )
        return
    
    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    chat_id_str = str(chat_id)
    
    # Перевіряємо, чи користувач адміністратор групи
    is_admin = await is_group_admin(user_id, chat_id, context)
    if not is_admin:
        await update.message.reply_text(
            "❌ Ви повинні бути адміністратором групи."
        )
        return
    
    # Перевіряємо, чи це той, хто активував
    if chat_id_str not in auto_report_config:
        await update.message.reply_text(
            "❌ Автоматична відправка для цієї групи ще не активована."
        )
        return
    
    if auto_report_config[chat_id_str]["admin_id"] != user_id:
        await update.message.reply_text(
            "❌ Деактивувати може тільки той адміністратор, який активував автоматичну відправку."
        )
        return
    
    del auto_report_config[chat_id_str]
    save_config()
    
    await update.message.reply_text(
        "✅ Автоматична відправка деактивована."
    )
    
    print(f"[LOG] /stoprecord деактивована для групи {chat_id}")


# === Рахунок повідомлень у групі ===
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print(f"[DEBUG] handle_message викликано для update: {update.update_id}")

    if not update.message:
        print("[DEBUG] update.message відсутній")
        return

    # Деталь: message.text буває None для фото/стікерів/анімацій, використовуємо caption (якщо є)
    msg_text = update.message.text or update.message.caption or ""

    # Визначаємо тип повідомлення
    msg_type = "unknown"
    if update.message.sticker:
        msg_type = "sticker"
    elif update.message.animation or (update.message.document and update.message.document.mime_type == "video/mp4"):
        msg_type = "gif"
    elif update.message.photo or (update.message.document and update.message.document.mime_type and update.message.document.mime_type.startswith("image/")):
        msg_type = "photo"
    elif msg_text.strip():  # перевіряємо, чи є текст (не тільки пробіли)
        msg_type = "text"

    print(f"Повідомлення надійшло: [{msg_type}] {msg_text or '<без тексту>'}")

    chat = update.effective_chat

    # ігноруємо особисті повідомлення
    if chat.type == "private":
        print(f"[DEBUG] Ігноруємо особисте повідомлення від {update.message.from_user.full_name}")
        return

    # ігноруємо команди (не враховуємо /start, /startrecord тощо як статистику)
    if update.message.text and update.message.text.startswith("/"):
        print(f"[DEBUG] Ігноруємо команду: {update.message.text}")
        return

    # лог для перевірки, що повідомлення надходять
    print(f"[LOG] Повідомлення від {update.message.from_user.full_name} в {chat.title}: [{msg_type}] {msg_text}")

    user = update.message.from_user
    chat_id = chat.id
    chat_names[chat_id] = chat.title

    user_id = user.id
    username = user.full_name

    month_key = datetime.now().strftime("%Y-%m")

    if chat_id not in user_stats:
        user_stats[chat_id] = {}
    if month_key not in user_stats[chat_id]:
        user_stats[chat_id][month_key] = {}
    if user_id not in user_stats[chat_id][month_key]:
        user_stats[chat_id][month_key][user_id] = {
            "name": username, 
            "text": 0, 
            "photo": 0, 
            "sticker": 0, 
            "gif": 0, 
            "total": 0
        }

    # Визначаємо тип повідомлення та збільшуємо відповідний лічильник
    message = update.message
    
    # Текстове повідомлення (не враховуємо captions у медіа)
    if message.text and not message.photo and not message.sticker and not message.animation and not message.document:
        user_stats[chat_id][month_key][user_id]["text"] += 1
        print(f"[STATS] +1 text для {username}")

    # Фото (також можливий в document якщо mime image/*)
    if message.photo or (message.document and message.document.mime_type and message.document.mime_type.startswith("image/")):
        user_stats[chat_id][month_key][user_id]["photo"] += 1
        print(f"[STATS] +1 photo для {username}")

    # Стикер
    if message.sticker:
        user_stats[chat_id][month_key][user_id]["sticker"] += 1
        print(f"[STATS] +1 sticker для {username}")

    # GIF анімація (animation + mp4 в document)
    if message.animation or (message.document and message.document.mime_type == "video/mp4"):
        user_stats[chat_id][month_key][user_id]["gif"] += 1
        print(f"[STATS] +1 gif для {username}")

    # Загальна кількість повідомлень
    user_stats[chat_id][month_key][user_id]["total"] += 1
    print(f"[STATS] Загалом: {user_stats[chat_id][month_key][user_id]}")
# === Перевірка, чи користувач адміністратор групи ===
async def is_group_admin(user_id: int, chat_id: int, context: ContextTypes.DEFAULT_TYPE) -> bool:
    try:
        member = await context.bot.get_chat_member(chat_id, user_id)
        # Користувач адміністратор, якщо він Creator або Administrator
        return member.status in ["creator", "administrator"]
    except Exception as e:
        print(f"[ERROR] Помилка при перевірці статусу адміністратора: {e}")
        return False

# === /groups в личке ===
async def groups_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        return

    if not chat_names:
        await update.message.reply_text("Я ще не бачив повідомлень у групах.")
        return

    user_id = update.effective_user.id
    
    # Фільтруємо групи, де користувач є адміністратором
    admin_groups = []
    for chat_id, name in chat_names.items():
        if await is_group_admin(user_id, chat_id, context):
            admin_groups.append((chat_id, name))
    
    if not admin_groups:
        await update.message.reply_text("Ви не є адміністратором жодної групи, де я записую статистику.")
        return

    # Створюємо inline кнопки для груп, де користувач адміністратор
    keyboard = []
    for chat_id, name in admin_groups:
        button = InlineKeyboardButton(text=name, callback_data=f"group_{chat_id}")
        keyboard.append([button])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Вибери групу (твої групи):"  , reply_markup=reply_markup)

# === Обробка нажаття на кнопку групи ===
async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not query.data.startswith("group_"):
        return
    
    user_id = query.from_user.id
    
    try:
        chat_id = int(query.data.replace("group_", ""))
        print(f"[DEBUG] Користувач {user_id} вибрав групу: {chat_id}")
    except ValueError as e:
        print(f"[ERROR] Помилка парсингу callback_data: {e}")
        await query.edit_message_text("Помилка команди")
        return
    
    # Перевіряємо, чи користувач адміністратор цієї групи
    is_admin = await is_group_admin(user_id, chat_id, context)
    if not is_admin:
        print(f"[ERROR] Користувач {user_id} не є адміністратором групи {chat_id}")
        await query.edit_message_text("❌ У вас немає доступу до статистики цієї групи. Ви повинні бути адміністратором.")
        return
    
    month_key = datetime.now().strftime("%Y-%m")
    print(f"[DEBUG] Поточний місяць: {month_key}, Наявні групи: {list(user_stats.keys())}")

    if chat_id not in user_stats:
        print(f"[ERROR] chat_id {chat_id} не знайдено в user_stats")
        await query.edit_message_text("Немає даних щодо цієї групи.")
        return
    
    if month_key not in user_stats[chat_id]:
        print(f"[ERROR] Місяць {month_key} не знайдено для {chat_id}. Наявні місяці: {list(user_stats[chat_id].keys())}")
        await query.edit_message_text("Немає даних щодо цієї групи.")
        return

    file_name = create_stats_workbook(chat_id, month_key)
    with open(file_name, "rb") as file:
        await query.message.reply_document(file)

# === Надсилання Excel файлу за групою ===
async def get_group_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        return

    user_id = update.effective_user.id
    command = update.message.text
    print(f"[DEBUG] Команда отримана: {command}")
    
    try:
        # Парсимо ID з команди /get_{chat_id}
        chat_id_str = command.replace("/get_", "").strip()
        chat_id = int(chat_id_str)
        print(f"[DEBUG] Розпарсений chat_id: {chat_id}")
    except ValueError as e:
        print(f"[ERROR] Помилка парсингу: {e}")
        await update.message.reply_text("Помилка команди")
        return
    
    # Перевіряємо, чи користувач адміністратор цієї групи
    is_admin = await is_group_admin(user_id, chat_id, context)
    if not is_admin:
        print(f"[ERROR] Користувач {user_id} не є адміністратором групи {chat_id}")
        await update.message.reply_text("❌ У вас немає доступу до статистики цієї групи. Ви повинні бути адміністратором.")
        return

    month_key = datetime.now().strftime("%Y-%m")
    print(f"[DEBUG] Поточний місяць: {month_key}, Наявні групи: {list(user_stats.keys())}")

    if chat_id not in user_stats:
        print(f"[ERROR] chat_id {chat_id} не знайдено в user_stats")
        await update.message.reply_text("Немає даних щодо цієї групи.")
        return
    
    if month_key not in user_stats[chat_id]:
        print(f"[ERROR] Місяць {month_key} не знайдено для {chat_id}. Наявні місяці: {list(user_stats[chat_id].keys())}")
        await update.message.reply_text("Немає даних щодо цієї групи.")
        return

    file_name = create_stats_workbook(chat_id, month_key)
    with open(file_name, "rb") as file:
        await update.message.reply_document(file)


# === Функція для генерування Excel для автоматичної відправки ===
def create_stats_workbook(chat_id: int, month_key: str) -> str:
    """Генерує Excel файл зі статистикою для групи. Повертає шлях до файлу."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stats"

    headers = ["Користувач", "Текст", "Фото", "Стикери", "GIF", "Загалом"]
    ws.append(headers)

    for user in user_stats[chat_id][month_key].values():
        ws.append([
            user["name"],
            user["text"],
            user["photo"],
            user["sticker"],
            user["gif"],
            user["total"]
        ])

    # Форматування
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            if cell.column != 1:
                cell.alignment = center

    # Розміри колонок
    widths = [30, 12, 12, 12, 12, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    file_name = f"stats_{chat_id}_{month_key}.xlsx"
    wb.save(file_name)
    return file_name


async def generate_excel_for_chat(chat_id: int) -> str:
    """Генерує Excel файл зі статистикою для групи. Повертає шлях до файлу або None"""
    month_key = (datetime.now() - timedelta(days=1)).strftime("%Y-%m")  # Попередній місяць
    
    if chat_id not in user_stats or month_key not in user_stats[chat_id]:
        print(f"[DEBUG] Немає даних для групи {chat_id} за місяць {month_key}")
        return None

    return create_stats_workbook(chat_id, month_key)


# === Функція для автоматичної відправки звітів ===
async def send_monthly_reports(app: Application):
    """Відправляє Excel звіти адміністраторам кожного місяця"""
    print("[LOG] Запущена функція send_monthly_reports")
    
    for chat_id_str, config in auto_report_config.items():
        try:
            if not config.get("enabled", False):
                continue
            
            chat_id = int(chat_id_str)
            admin_id = config["admin_id"]
            admin_name = config["admin_name"]
            chat_title = config.get("chat_title", "Невідома група")
            
            # Генеруємо Excel
            file_path = await generate_excel_for_chat(chat_id)
            
            if file_path and os.path.exists(file_path):
                # Відправляємо адміністратору в особистій повідмленні
                try:
                    with open(file_path, "rb") as file:
                        month_key = (datetime.now() - timedelta(days=1)).strftime("%B %Y")
                        await app.bot.send_document(
                            chat_id=admin_id,
                            document=file,
                            caption=f"📊 Статистика групи '{chat_title}' за {month_key}"
                        )
                    print(f"[LOG] Excel звіт відправлено адміністратору {admin_id} для групи {chat_id}")
                    
                    # Видаляємо файл після відправки
                    os.remove(file_path)
                except Exception as e:
                    print(f"[ERROR] Помилка при відправці файлу адміністратору {admin_id}: {e}")
            else:
                print(f"[DEBUG] Немає даних для відправки групі {chat_id}")
        
        except Exception as e:
            print(f"[ERROR] Помилка при обробці групи {chat_id_str}: {e}")


# === Запуск бота ===
def main():
    print("[STARTUP] Ініціалізація бота...")
    app = Application.builder().token(TOKEN).build()
    print("[STARTUP] Application створено")
    
    # Завантажуємо конфігурацію при старті
    load_config()
    print("[STARTUP] Конфігурація завантажена")
    
    # Встановлюємо scheduler для автоматичної відправки звітів
    scheduler = AsyncIOScheduler()
    print("[STARTUP] Scheduler створено")
    
    # Додаємо job для відправки звітів першого числа кожного місяця о 9:00
    scheduler.add_job(
        send_monthly_reports,
        CronTrigger(day=1, hour=9, minute=0),
        args=[app],
        id='send_monthly_reports',
        name='Send Monthly Reports',
        replace_existing=True
    )
    print("[STARTUP] Job додано до scheduler")
    
    # Асинхронна функція для ініціалізації scheduler
    async def init_scheduler(application):
        scheduler.start()
        print("[STARTUP] Scheduler запущено")
    
    app.post_init = init_scheduler

    # команди
    print("[STARTUP] Реєстрація команд...")
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("groups", groups_command))
    app.add_handler(CommandHandler("startrecord", startrecord_command))
    app.add_handler(CommandHandler("stoprecord", stoprecord_command))
    app.add_handler(CallbackQueryHandler(button_callback))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/get_-?\d+$"), get_group_stats))

    # обробка всіх повідомлень у групах (включно фото/стікери/animaції)
    app.add_handler(MessageHandler(filters.ALL, handle_message))
    print("[STARTUP] Всі handlers зареєстровано")

    print("Бот запущен...")
    app.run_polling()


if __name__ == "__main__":
    main()